import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

def add_info_to_excel(excel_path, all_pdfs_info): 

    # column details
    Proposal_Number_column = 'A'
    title_column = 'B'
    title_length_column = 'C'
    total_file_size_column = 'D'
    pi_column = "E"
    submitted_by_column = "F"
    date_column = "G"
    phd_answer_column = "W"
    early_career_column = "V"


    # load in excel workbook
    wb = load_workbook(excel_path)

    # specify which excel sheet we're using
    sheet_name = 'NEW PROPOSALS'  # Replace with your sheet name
    sheet = wb[sheet_name]

    # read in all pdf info and insert into excel sheet
    start_row = 8
    for i in range(len(all_pdfs_info)): 
        
        for key,value in all_pdfs_info[i].items():

            cell_coord = None

            # add all info to excel
            if key == "proposal number":
                cell_coord = str(Proposal_Number_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "title":
                cell_coord = str(title_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "title length":
                cell_coord = str(title_length_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "file size":
                cell_coord = str(total_file_size_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "PI": 
                cell_coord = str(pi_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "date":
                cell_coord = str(date_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "phd":
                cell_coord = str(phd_answer_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value
            
            elif key == "EarlyTrack": 
                cell_coord = str(early_career_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "submitted by": 
                cell_coord = str(submitted_by_column) + str(start_row+i)
                cell = sheet[cell_coord]
                cell.value = value

            elif key == "pi_submitter_alert": 
                if value == True:
                    cell_coord = str(submitted_by_column) + str(start_row+i)
                    fill_color = 'ffffa1'
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    cell = sheet[cell_coord]
                    cell.fill = fill

            else:
                print(f"ERROR: unable to add into to excel. {key}: {value}")
                pass

            # make sure text isn't bold
            font = Font(bold=False)
            cell.font = font

    wb.save(excel_path)

    

















#df = pd.read_excel(excel_path)

#   # create a copy of the excel template
#   #writer = pd.ExcelWriter(excel_path, engine='xlsxwriter')

#   # Specify the row number you want to read (0-based indexing)
#   row_number = 2

#   # Get the row as a Series
#   row_data = df.loc[row_number]

#   # Access the values in the row
#   values = row_data.values

#   # Convert values to a list if needed
#   values_list = values.tolist()

#   # Print the row values
#   print(values_list)

#   # open the newly created empty excel

  #print(excel_path)
